#trial
from psychopy import visual, core, event, monitors
import common, time, config, random
from openpyxl import Workbook, load_workbook

def call_frame(call_name, frame_name, window):
    frame_dict = {
        "Profession1":  visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos =  [-0.25, 0.2], height = 0.07, wrapWidth=None),
        "Profession2":  visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos =  [0.25, 0.2], height = 0.07, wrapWidth=None),
        "Attribute":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0,0], height = 0.07, wrapWidth=None),
        "Baserate1":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [-0.25, -0.2], height = 0.07, wrapWidth=None), 
        "Baserate2":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0.25, -0.2], height = 0.07, wrapWidth=None),
        "Answeroption1":visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [-0.3, 0], height = 0.07, wrapWidth=None),
        "Answeroption2":visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0.3, 0], height = 0.07, wrapWidth=None)
        }
    #if we call the function with the right name, we return the visual stimuli
    if frame_name in frame_dict:
        return frame_dict[frame_name]

def main_loop(dict_values, workbook_output, window, excel_number, tk):
    timer = core.Clock()

    tk.sendMessage("trial number: " + str(excel_number-1))
            
    prof_1 = dict_values["Profession1"]
    prof_2 = dict_values["Profession2"]
    attribute = dict_values["Attribute"]
    base_rate1 = dict_values["Baserate1"]
    base_rate2 = dict_values["Baserate2"]
    answer_1 = dict_values["Answeroption1"]
    answer_2 = dict_values["Answeroption2"]
    
    time.sleep(2)
    
    tk.sendMessage('this_study_contains')

    common.fixationCross("small", 1, 0, window)

    frame1_studyContains = visual.TextStim(window, text = u"This study contains:", font = 'Arial', units = 'norm', pos = [0, 0.5], height = 0.07, wrapWidth=None)
    frame_1_profession1 = call_frame(prof_1, "Profession1", window)
    frame_1_profession2 = call_frame(prof_2, "Profession2", window)
    
    frames_list = [frame1_studyContains, frame_1_profession1, frame_1_profession2]
    for frame in frames_list:
        frame.draw()

    window.flip()
    time.sleep(1.8)

    common.fixationCross("small", 1, 0, window)
    
    tk.sendMessage('Attribute')

    frame_2_attribute = call_frame(attribute, "Attribute", window)
    frame_2_attribute.draw()
    window.flip()
    time.sleep(1.8)
	
    common.fixationCross("small", 1, 0, window)
	
    tk.sendMessage('Baserates/Answers')
    
    frame_3_baserate1 = call_frame(base_rate1, "Baserate1", window)
    frame_3_baserate2 = call_frame(base_rate2, "Baserate2", window)
    
    frames_list = [frame_3_baserate1, frame_3_baserate2, frame_1_profession1, frame_1_profession2]
    for frame in frames_list:
        frame.draw()

    window.flip()
    
    timer.reset()
    startTime = timer.getTime()
    event.clearEvents() #if one accidentaly clicks a key before the selection process, it will not register
    while (startTime < 4):
        keys = event.getKeys()
        if (config.answer_left) in keys:
            
            workbook_output["E"+str(excel_number)] = "Answeroption1"
            workbook_output["F"+str(excel_number)] = startTime
            break
        
        elif (config.answer_right) in keys:
            
            workbook_output["E"+str(excel_number)] = "Answeroption2"
            workbook_output["F"+str(excel_number)] = startTime
            break

        core.wait(0.1)
        startTime += 0.1
        
    tk.sendMessage("Selected answer")
    #if time surpased the limit, we add a no response to the file
    if startTime > 4:
        workbook_output["E"+str(excel_number)] = "No response"
        workbook_output["F"+str(excel_number)] = 4
        
    window.flip()
    time.sleep(2)
    
"""

go through excel file, based on the numbers that are drawn from the list of all variables
"""
def draw_from_number(number, list, workbook):
    value_dict = {"Trial": "", "Condition": "", "Profession1": "", "Profession2": "", "Attribute": "", 
    "Baserate1": "", "Baserate2": "", "Answeroption1": "", "Answeroption2": "",
    "Correct": ""} 
    
    for column in range (1, 11): #1-10, Trial - > Correct
        
        row_name = workbook.cell(row = 1, column = column) #Condition - > Correct answer
        cell_name = workbook.cell(row = number, column = column) #Americans -> Answeroption1/2
        value_dict[row_name.value] = cell_name.value #set value from excel file to dictionary
    
    list.append(value_dict)
    
def runTrial(workbook_input, window, dest_filename, tk):
    
	
    list_N = [] #8-29 in excel sheet
    list_C = [] #30-73 in excel sheet
    list_I = [] #74-139 in excel sheet
    
    #Append excel numbers to seperate lists in python
    for i in range(8, 30): list_N.append(i)
    for i in range(30, 74): list_C.append(i)
    for i in range(74, 140): list_I.append(i)
    
    #main loop
    num = 0 #Checking if it goes through all trials
    excel_number = 2 #jump one ahead of the titles
    last_list = [] #used for checking multiple hits in a row
	
    for i in range(1, 23): #23 = All
    
        list_dict = []
        
        #add 6 draws to the list
        num_N = random.choice(list_N)
        draw_from_number(num_N, list_dict, workbook_input)
        list_N.remove(num_N)
        
        num_C1 = random.choice(list_C)
        draw_from_number(num_C1, list_dict, workbook_input)
        list_C.remove(num_C1)
        
        num_C2 = random.choice(list_C)
        draw_from_number(num_C2, list_dict, workbook_input)
        list_C.remove(num_C2)
      
        num_I1 = random.choice(list_I)
        draw_from_number(num_I1, list_dict, workbook_input)
        list_I.remove(num_I1)
        
        num_I2 = random.choice(list_I)
        draw_from_number(num_I2, list_dict, workbook_input)
        list_I.remove(num_I2)
        
        num_I3 = random.choice(list_I)
        draw_from_number(num_I3, list_dict, workbook_input)
        list_I.remove(num_I3)

        random.shuffle(list_dict)
        
        #Avoid 4 in a row, right now maximum is 3 in a row
        if last_list:
            if last_list[0] == last_list[1]:
                while (last_list[1] == list_dict[0]["Condition"]):
                    random.shuffle(list_dict)
                          
        last_list = []
        #Go through sub_list of 6 draws            
        #for k in range(0, 2): #FOR TESTING
        for k in range(0, (len(list_dict))): #Original
            
             #append the last two
            if k > 3:
                last_list.append(list_dict[k]["Condition"])
            
            #Activate workbook for each iteration
            wb = load_workbook(filename = dest_filename)
            output_file = wb.active

            trial_number = (int(list_dict[k]["Trial"])) + 1

            output_file['C'+ str(excel_number)] = trial_number
            output_file['D'+ str(excel_number)] = list_dict[k]["Condition"]

            main_loop(list_dict[k], output_file, window, excel_number, tk)
            
            excel_number += 1
            num += 1 
            			
            #save output workbook here, saves each trial
            wb.save(filename = dest_filename)
            
    #current values is 132 trials

    