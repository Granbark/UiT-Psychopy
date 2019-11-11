from psychopy import visual, core, event
import config, common, random, os, time
from openpyxl import Workbook, load_workbook
import sounddevice as sd
import soundfile as sf

def call_frame(call_name, frame_name, window):
    frame_dict = {
        "Profession1":  visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos =  [-0.25, 0.2], height = 0.07, wrapWidth=None),
        "Profession2":  visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos =  [0.25, 0.2], height = 0.07, wrapWidth=None),
        "Attribute":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0,0], height = 0.07, wrapWidth=None),
        "Baserate1":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [-0.25, -0.2], height = 0.07, wrapWidth=None), 
        "Baserate2":    visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0.25, -0.2], height = 0.07, wrapWidth=None),
        "Answeroption1":visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [-0.3, 0], height = 0.07, wrapWidth=None),
        "Answeroption2":visual.TextStim(window, text = call_name, font = 'Arial', units = 'norm', pos = [0.3, 0], height = 0.07, wrapWidth=None)
        }
    #if we call the function with the right name, we return the visual stimuli
    if frame_name in frame_dict:
        return frame_dict[frame_name]

def draw_from_number(number, list, workbook):
    value_dict = {"Trial": "", "Condition": "", "Profession1": "", "Profession2": "", "Attribute": "", 
    "Baserate1": "", "Baserate2": "", "Answeroption1": "", "Answeroption2": "",
    "Correct": "", "wav_profession": ""} 
    
    for column in range (1, 12): #1-11, Trial - > Correct
        
        row_name = workbook.cell(row = 1, column = column) #Condition - > Correct answer
        cell_name = workbook.cell(row = number, column = column) #Americans -> Answeroption1/2
        value_dict[row_name.value] = cell_name.value #set value from excel file to dictionary
    
    list.append(value_dict)

def mainLoop(dict_values, window, audioFolder, audioDevice, excelNumber, output_file, tk):
    timer = core.Clock()

    tk.sendMessage("trial number: " + str(excelNumber-1))

    prof_1 = dict_values["Profession1"]
    prof_2 = dict_values["Profession2"]
    attribute = dict_values["Attribute"]
    base_rate1 = dict_values["Baserate1"]
    base_rate2 = dict_values["Baserate2"]

    likely_sound, likely_fs = sf.read("is_this_person_more_likely_a.wav")
    attribute_sound = attribute + ".wav"
    profession_sound = dict_values["wav_profession"]
	
    original_folder = os.getcwd()
    
    os.chdir(audioFolder)
    attribute_sound, attribute_fs = sf.read(attribute_sound)
    profession_sound, profession_fs = sf.read(profession_sound)
    
    os.chdir(original_folder)

    window.flip()
    time.sleep(0.5)

    common.fixationCross("large", 0, 0.5, window)

    tk.sendMessage('this_study_contains')

    frame1_studyContains = visual.TextStim(window, text = u"This study contains:", font = 'Arial', units = 'norm', pos = [0, 0.5], height = 0.07, wrapWidth=None)
    frame_1_profession1 = call_frame(prof_1, "Profession1", window)
    frame_1_profession2 = call_frame(prof_2, "Profession2", window)
    frames_list = [frame1_studyContains, frame_1_profession1, frame_1_profession2]
    
    for frame in frames_list:
        frame.draw()

    window.flip()
    time.sleep(1.8)

    common.fixationCross("large", 1, 0, window)

    tk.sendMessage('base_rate')
	
    frame_3_baserate1 = call_frame(base_rate1, "Baserate1", window)
    frame_3_baserate2 = call_frame(base_rate2, "Baserate2", window)
    frames_list = [frame_3_baserate1, frame_3_baserate2, frame_1_profession1, frame_1_profession2]

    for frame in frames_list:
        frame.draw()

    window.flip()
    time.sleep(3.6)

    tk.sendMessage('attribute_start')
	
    sd.play(attribute_sound, attribute_fs, device = audioDevice)

    common.fixationCross("large", 0, 3, window)

    tk.sendMessage('attribute_end')

    window.flip()
    time.sleep(0.2)

    tk.sendMessage('is_this_person_more_likely_a')
	
    sd.play(likely_sound, likely_fs, device = audioDevice) # lasts for two seconds

    common.fixationCross("large", 0, 3, window)
    sd.wait()
    sd.play(profession_sound, profession_fs, device = audioDevice)
    time.sleep(1.2)
    
    tk.sendMessage('Answer_start')

    timer.reset()
    startTime = timer.getTime()
    event.clearEvents() #if one accidentaly clicks a key before the selection process, it will not register
    
    while (startTime < 4):
        keys = event.getKeys()
        if (config.answer_left) in keys:
            
            output_file["E"+str(excelNumber)] = "Answeroption1"
            output_file["F"+str(excelNumber)] = startTime
            break
        elif (config.answer_right) in keys:
            
            output_file["E"+str(excelNumber)] = "Answeroption2"
            output_file["F"+str(excelNumber)] = startTime
            break

        core.wait(0.1)
        startTime += 0.1
        
    #if time surpased the limit, we add a no response to the file
    if startTime > 4:
        output_file["E"+str(excelNumber)] = "No response"
        output_file["F"+str(excelNumber)] = 4
        return
    
    tk.sendMessage('Answer_end')
	
    time_jitt = random.uniform(3,5)
    common.fixationCross("large", 0, time_jitt, window)
    window.flip()
    time.sleep(random.uniform(1,2.5))

def runTrial(baseRates, window, subjectExcel, eyeTracker, audioDevice):

    audioFolder = os.getcwd() + "\\audio_files"

    list_N = [] #1-22 in excel sheet
    list_C = [] #30-67 in excel sheet
    list_I = [] #67-132 in excel sheet
        
    #This has to be adjusted for each individual excel file
    #Append excel numbers to seperate lists in python
    for i in range(4, 26): list_N.append(i) 
    for i in range(26, 70): list_C.append(i) 
    for i in range(70, 136): list_I.append(i) 
     
    #main loop
    num = 0 #Checking if it goes through all trials
    excel_number = 2 #jump one ahead of the titles
    last_list = [] #used for checking multiple hits in a row
    for i in range(1, 23):
        #add 6 draws to the list
        
        list_dict = []
        
        num_N = random.choice(list_N)
        draw_from_number(num_N, list_dict, baseRates)
        list_N.remove(num_N)
        
        num_C1 = random.choice(list_C)
        draw_from_number(num_C1, list_dict, baseRates)
        list_C.remove(num_C1)
        
        num_C2 = random.choice(list_C)
        draw_from_number(num_C2, list_dict, baseRates)
        list_C.remove(num_C2)
      
        num_I1 = random.choice(list_I)
        draw_from_number(num_I1, list_dict, baseRates)
        list_I.remove(num_I1)
        
        num_I2 = random.choice(list_I)
        draw_from_number(num_I2, list_dict, baseRates)
        list_I.remove(num_I2)
        
        num_I3 = random.choice(list_I)
        draw_from_number(num_I3, list_dict, baseRates)
        list_I.remove(num_I3)

        random.shuffle(list_dict)
        
        #Avoid 4 in a row, right now maximum is 3 in a row
        if last_list:
            if last_list[0] == last_list[1]:
                while (last_list[1] == list_dict[0]["Condition"]):
                    random.shuffle(list_dict)
                          
        last_list = []

        #Go through sub_list of 6 draws     
        #for k in range(0,2):
        for k in range(0, (len(list_dict))):
            
            #append the last two, used for duplicate check
            if k > 3:
                last_list.append(list_dict[k]["Condition"])
                
            #activate it for each iteration
            wb = load_workbook(filename = subjectExcel)
            output_file = wb.active

            trial_number = (int(list_dict[k]["Trial"])) + 1

            output_file['C'+ str(excel_number)] = trial_number
            output_file['D'+ str(excel_number)] = list_dict[k]["Condition"]

            #excel number here
            mainLoop(list_dict[k], window, audioFolder, audioDevice, excel_number, output_file, eyeTracker)

            num += 1 
            excel_number += 1

            #save output workbook here, saves each trial
            
            wb.save(filename = subjectExcel)