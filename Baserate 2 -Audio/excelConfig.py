from openpyxl import Workbook, load_workbook
from psychopy import data

def excelBaseRates():

    base_rates = load_workbook(filename = '20190827_Base_rate_exp2.xlsx')
    base_rates = base_rates.active
    return base_rates

def excelNewSubject(expInfo):
    
    wb = Workbook()
    date = data.getDateStr()
    
    dest_filename = expInfo['SubjectNO'] + "-" + expInfo['SubjectInitials'] + "-" + date + ".xlsx"
    
    workbook_results = wb.active
    workbook_results.title = "Information"

    #Trial number (consecutive numbering from 1 to # of trials in the excel file), 
    #Trial type (column condition), 
    #ResponseType (i.e. AnswerOption1/2), 
    #ResponseTime 

    workbook_results['A1'] = u"ID"
    workbook_results['B1'] = u"Date"
    workbook_results['A2'] = expInfo['SubjectNO']
    workbook_results['B2'] = date

    workbook_results['C1'] = u"Trial Number"
    workbook_results['D1'] = u"Trial Type"
    workbook_results['E1'] = u"Response Type"
    workbook_results['F1'] = u"Response Time"

    wb.save(filename = dest_filename)

    return dest_filename








